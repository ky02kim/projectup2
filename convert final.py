
"""
기업건강의학센터 업무실적 대시보드 — 엑셀 변환기 (통합본)
=====================================================
사용법:
  1. 이 파일(convert.py 또는 convert.exe)을 다음 파일/폴더와 같은 위치에 놓으세요.
       - '업무 실적.xlsx'                         (기존 KNX/자료요청서 기능)
       - '거래처별 사후관리 집계_YYYYMMDDHHMMSS.xlsx' 여러 개  (신규 기능)
  2. 더블클릭 실행
       → data/knx.js, data/req.js           (기존 기능, 변경 없음)
       → data/hgc.js                        (신규 기능)
  3. 생성된 data/ 폴더를 index.html과 함께 NAS에 복사

────────────────────────────────────────────────────────
[기존] 업무 실적.xlsx 구조
  시트1 (KNX 발송)
    A: 발송일자  C: 사업장명  D: 예/종건  E: 거래처구분
    H~P: 9개 항목  Q: 비고  R: 발송인  S: 제공형태
  시트2 (자료요청서)
    C: 발송일자  E: 사업장명  F: 예/종건  G: 거래처구분
    K~S: 9개 항목  T: 비고  U: 발송인

[신규] 거래처별 사후관리 집계_YYYYMMDDHHMMSS.xlsx 구조
  A2: "건진센터 : 종건(U)" 또는 "건진센터 : 예건(G)"  ← 종건/예건 구분
  A3: "건진기간 : YYYY-MM-DD ~ YYYY-MM-DD"           ← 건진기간
  헤더행(자동탐색, E열에 '거래처명' 텍스트가 있는 행): E, N, V, AB~AE
    E : 거래처명 (string)
    N : 노동부보고 (Y/N)
    V : 접수 (int)
    AB~AE : 4개 항목 (Y/N, 항목명은 헤더행에서 자동 추출)
  필터: N == 'Y'  AND  V >= 1  인 행만 사용

  같은 '건진기간' + 같은 '종건/예건' 그룹 내에서, 파일명의 타임스탬프 기준으로
    - 가장 오래된 파일  → 발송대상
    - 가장 최신 파일    → 발송완료
  (파일이 1개뿐인 그룹은 발송대상=발송완료로 동일 처리, 콘솔에 경고 출력)
"""

import os
import re
import sys
import glob
import json
import traceback
from datetime import datetime, date


# ── openpyxl 임포트 ─────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("=" * 50)
    print("오류: openpyxl이 설치되어 있지 않습니다.")
    print("cmd에서 실행: pip install openpyxl")
    print("=" * 50)
    input("엔터를 누르면 종료합니다...")
    sys.exit(1)

# ── 설정 ────────────────────────────────────────────────

EXCEL_FILE = "[기업건강연구소]업무 실적_2026년~.xlsx"       # 기존 기능 엑셀 파일명 (고정)
HGC_PATTERN = "*사후관리*집계*.xlsx"  # 신규 기능 파일명 패턴
OUTPUT_DIR = "data"                 # 출력 폴더

ITEM_LABELS = [
    '사후세로형', '사후가로형', '뇌심', '직무', '정신',
    '동의자결과', '사업장양식', '사이트업로드', '통계자료'
]


# ── 공통 헬퍼 ───────────────────────────────────────────
def get_script_dir():
    """실행 파일(exe 포함)의 실제 위치 반환"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def safe_str(cell):
    """셀 값을 안전하게 문자열로 변환"""
    if cell is None:
        return ''
    v = cell.value if hasattr(cell, 'value') else cell
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return normalize_date(v) or ''
    return str(v).strip()


def normalize_date(v):
    """날짜 값을 YYYY-MM-DD 문자열로 정규화"""
    if v is None:
        return None

    if isinstance(v, (datetime, date)):
        try:
            y = v.year if hasattr(v, 'year') else None
            m = v.month if hasattr(v, 'month') else None
            d = v.day if hasattr(v, 'day') else None
            if y and 2020 <= y <= 2050:
                return f"{y:04d}-{m:02d}-{d:02d}"
        except Exception:
            return None
        return None

    s = str(v).strip()
    if not s:
        return None

    s = s.replace('/', '-').replace('.', '-')

    if len(s) == 8 and s.isdigit():
        s = f"{s[:4]}-{s[4:6]}-{s[6:8]}"

    if len(s) >= 10:
        try:
            parts = s[:10].split('-')
            if len(parts) == 3:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                if 2020 <= y <= 2050 and 1 <= m <= 12 and 1 <= d <= 31:
                    return f"{y:04d}-{m:02d}-{d:02d}"
        except Exception:
            pass

    return None


def to_check(cell):
    return 1 if safe_str(cell) == "1" else 0


def to_bool_yn(cell):
    """Y/N 형태의 셀을 True/False로 변환"""
    v = safe_str(cell).strip().upper()
    return v == 'Y'


def to_int_safe(cell):
    """정수로 안전 변환 (실패 시 0)"""
    v = cell.value if hasattr(cell, 'value') else cell
    if v is None or v == '':
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def is_empty_row(row, min_col, max_col):
    """지정 범위 셀이 모두 비어있으면 True"""
    for c in range(min_col, max_col + 1):
        if row[c - 1].value not in (None, ''):
            return False
    return True


# ── KNX 시트 파싱 (기존, 변경 없음) ──────────────────────
def parse_knx(ws):
    result = []
    rows = list(ws.iter_rows(min_row=2))

    for row in rows:
        if len(row) < 1:
            continue
        date_val = normalize_date(row[0].value)
        if not date_val:
            continue

        parts = date_val.split('-')
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        if not (2020 <= y <= 2050 and 1 <= m <= 12):
            continue

        items = []
        for k in range(9):
            idx = 7 + k
            cell = row[idx] if idx < len(row) else None
            items.append(to_check(cell) if cell else 0)

        def g(idx):
            return safe_str(row[idx]) if idx < len(row) else ''

        result.append({
            'date':      date_val,
            'y':         y,
            'm':         m,
            'd':         d,
            'workplace': g(2),
            'jongYe':    g(3),
            'group':     g(4),
            'items':     items,
            'itemSum':   sum(items),
            'memo':      g(16),
            'staff':     g(17),
            'provType':  g(18),
        })

    return result


# ── 자료요청서 시트 파싱 (기존, 변경 없음) ───────────────
def parse_req(ws):
    result = []
    rows = list(ws.iter_rows(min_row=2))

    for row in rows:
        if len(row) < 3:
            continue
        date_val = normalize_date(row[2].value)
        if not date_val:
            continue

        parts = date_val.split('-')
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        if not (2020 <= y <= 2050 and 1 <= m <= 12):
            continue

        items = []
        for k in range(9):
            idx = 10 + k
            cell = row[idx] if idx < len(row) else None
            items.append(to_check(cell) if cell else 0)

        def g(idx):
            return safe_str(row[idx]) if idx < len(row) else ''

        result.append({
            'date':      date_val,
            'y':         y,
            'm':         m,
            'd':         d,
            'workplace': g(4),
            'jongYe':    g(5),
            'group':     g(6),
            'items':     items,
            'itemSum':   sum(items),
            'note':      g(19),
            'sender':    g(20),
        })

    return result


# ══════════════════════════════════════════════════════
#  신규 기능: 거래처별 사후관리 집계 전처리
# ══════════════════════════════════════════════════════

HGC_COL = {
    'E': 5, 'N': 14, 'V': 22,
    'AB': 28, 'AC': 29, 'AD': 30, 'AE': 31,
}

TS_RE = re.compile(r'_(\d{14})\.xlsx$', re.IGNORECASE)
PERIOD_RE = re.compile(
    r'(\d{4}[-./]\d{1,2}[-./]\d{1,2})\s*~\s*(\d{4}[-./]\d{1,2}[-./]\d{1,2})'
)


def find_hgc_files(base_dir):
    """폴더 내 '거래처별 사후관리 집계_*.xlsx' 파일 목록 + 타임스탬프 추출"""
    paths = glob.glob(os.path.join(base_dir, HGC_PATTERN))
    files = []
    for p in paths:
        fname = os.path.basename(p)
        m = TS_RE.search(fname)
        if not m:
            print(f"       [경고] 파일명에서 타임스탬프를 못 찾음 (건너뜀): {fname}")
            continue
        ts_str = m.group(1)
        try:
            ts = datetime.strptime(ts_str, '%Y%m%d%H%M%S')
        except ValueError:
            print(f"       [경고] 타임스탬프 파싱 실패 (건너뜀): {fname}")
            continue
        files.append({'path': p, 'fname': fname, 'timestamp': ts})
    return files


def parse_hgc_meta(ws):
    """A2(종건/예건), A3(건진기간) 파싱"""
    a2 = safe_str(ws['A2'])
    a3 = safe_str(ws['A3'])

    if '종건' in a2:
        jong_ye = '종건'
    elif '예건' in a2:
        jong_ye = '예건'
    else:
        jong_ye = None

    period_start, period_end = None, None
    m = PERIOD_RE.search(a3)
    if m:
        period_start = normalize_date(m.group(1))
        period_end = normalize_date(m.group(2))

    return jong_ye, period_start, period_end


HGC_CONSECUTIVE_BLANK_LIMIT = 30   # 이만큼 연속 빈 행이면 데이터 끝으로 간주
HGC_HARD_ROW_CAP = 50000           # 만약을 대비한 절대 상한 (경고 후 중단)




def parse_hgc_data(ws):
    """
    한 파일의 데이터 파싱.
    반환: (item_labels[4], rows[])
      rows[i] = {'거래처명':str, '노동부보고':bool, '접수':int, 'items':[bool*4]}
    필터: 노동부보고 == 'Y' and 접수 >= 1
    """
    header_row = 6
   
    e_i, n_i, v_i = HGC_COL['E'] - 1, HGC_COL['N'] - 1, HGC_COL['V'] - 1
    ab_i, ac_i, ad_i, ae_i = (HGC_COL[c] - 1 for c in ('AB', 'AC', 'AD', 'AE'))
    max_col = HGC_COL['AE']

    header_vals = None
    rows = []
    consecutive_blank = 0
    scanned = 0

    # 성능/안정성 핵심: iter_rows()로 한 번만 스트리밍 스캔한다.
    # per-cell ws.cell(row, col) 랜덤 접근은 read_only 워크북에서 매우 느려서,
    # 서식이 수만~수십만 행까지 남아있는 실제 업무 파일에서 응답이 멈춘 것처럼
    # 보이는 원인이었음 (테스트 결과 동일 조건에서 1000배 이상 속도 차이).
    for row in ws.iter_rows(min_row=header_row, max_col=max_col, values_only=True):
        if header_vals is None:
            header_vals = row
            continue

        scanned += 1
        if scanned > HGC_HARD_ROW_CAP:
            print(f"       [경고] 데이터 스캔이 {HGC_HARD_ROW_CAP}행을 넘어 강제 중단했습니다. "
                  f"파일 구조를 확인해주세요.")
            break

        name = row[e_i] if e_i < len(row) else None
        v_raw = row[v_i] if v_i < len(row) else None
        name_str = '' if name is None else str(name).strip()
        v_str = '' if v_raw is None else str(v_raw).strip()

        if not name_str and not v_str:
            consecutive_blank += 1
            if consecutive_blank >= HGC_CONSECUTIVE_BLANK_LIMIT:
                break
            continue
        consecutive_blank = 0

        n_raw = row[n_i] if n_i < len(row) else None
        labor_report = str(n_raw).strip().upper() == 'Y' if n_raw is not None else False
        try:
            jeopsu = int(float(v_raw)) if v_raw not in (None, '') else 0
        except (ValueError, TypeError):
            jeopsu = 0

        if labor_report and jeopsu >= 1:
            def yn(idx):
                val = row[idx] if idx < len(row) else None
                return str(val).strip().upper() == 'Y' if val is not None else False

            rows.append({
                '거래처명': name_str,
                '노동부보고': labor_report,
                '접수': jeopsu,
                'items': [yn(ab_i), yn(ac_i), yn(ad_i), yn(ae_i)],
            })

    item_labels = [
        (str(header_vals[HGC_COL[c] - 1]).strip()
         if header_vals and HGC_COL[c] - 1 < len(header_vals) and header_vals[HGC_COL[c] - 1]
         else c)
        for c in ('AB', 'AC', 'AD', 'AE')
    ]

    return item_labels, rows


def load_hgc_files(base_dir):
    """폴더 내 모든 대상 파일을 열어 메타+데이터 추출"""
    files = find_hgc_files(base_dir)
    print(f"       → 대상 파일 {len(files)}개 발견")
    parsed = []
    for i, f in enumerate(files, 1):
        print(f"       [{i}/{len(files)}] 처리 중: {f['fname']}")
        try:
            wb = openpyxl.load_workbook(f['path'], data_only=True, read_only=True)
            ws = wb.active

            jong_ye, period_start, period_end = parse_hgc_meta(ws)
            if jong_ye is None or period_start is None or period_end is None:
                print("           [경고] A2/A3 메타 파싱 실패 (건너뜀)")
                wb.close()
                continue

            item_labels, rows = parse_hgc_data(ws)
            wb.close()

            parsed.append({
                'fname':        f['fname'],
                'timestamp':    f['timestamp'],
                'jongYe':       jong_ye,
                'periodStart':  period_start,
                'periodEnd':    period_end,
                'itemLabels':   item_labels,
                'rows':         rows,
            })
            print(f"           → {jong_ye}, {period_start}~{period_end}, {len(rows)}건")
        except Exception as e:
            print(f"           [오류] 처리 실패 (건너뜀): {e}")
            continue
    return parsed


def group_hgc(parsed_files):
    """
    같은 기간 + 거래처 + 종건/예건 기준으로 묶어
    가장 오래된 파일 = target
    가장 최신 파일 = complete
    """

    groups = {}

    # 행(Row) 단위로 그룹 생성
    for pf in parsed_files:
        for row in pf["rows"]:

            key = (
                pf["periodStart"],
                pf["periodEnd"],
                pf["jongYe"],
                row["거래처명"]
            )

            groups.setdefault(key, []).append({
                "timestamp": pf["timestamp"],
                "periodStart": pf["periodStart"],
                "periodEnd": pf["periodEnd"],
                "jongYe": pf["jongYe"],
                "workplace": row["거래처명"],
                "receive": row["접수"],
                "report": row["노동부보고"],
                "items": row["items"]
            })

    result = []

    for rows in groups.values():

        rows.sort(key=lambda x: x["timestamp"])

        # 파일이 하나면 target만 생성
        if len(rows) == 1:
            process_list = [("target", rows[0])]
        else:
            process_list = [
                ("target", rows[0]),
                ("complete", rows[-1])
            ]

        for status, r in process_list:

            result.append({
                "status": status,
                "fileDate": r["timestamp"].strftime("%Y-%m-%d"),
                "fileTimestamp": r["timestamp"].strftime("%Y-%m-%d %H:%M:%S"),
                "periodStart": r["periodStart"],
                "periodEnd": r["periodEnd"],
                "jongYe": r["jongYe"],
                "workplace": r["workplace"],
                "receive": r["receive"],
                "report": r["report"],
                "items": r["items"]
            })

    result.sort(key=lambda x: (
        x["periodStart"],
        x["jongYe"],
        x["workplace"],
        x["status"]
    ))

    return result
    # 건진기간 시작일 -> 종건/예건 순으로 정렬
    result.sort(key=lambda x: (x['periodStart'], x['jongYe']))
    return result


def process_hgc(base_dir):
    """신규 기능 전체 파이프라인 실행"""
    parsed = load_hgc_files(base_dir)
    if not parsed:
        print("       [정보] 대상 파일이 없거나 모두 파싱 실패했습니다.")
        return []
    return group_hgc(parsed)


# ── JS 파일 출력 ─────────────────────────────────────────
def write_js(filepath, var_name, data, generated_at):
    json_str = json.dumps(data, ensure_ascii=False, indent=2)
    content = f"""/* 자동 생성 파일 — convert.py
   생성 시각: {generated_at}
   ※ 직접 수정하지 마세요. 엑셀 수정 후 convert.exe를 다시 실행하세요. */
window.{var_name} = {json_str};
"""
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)


# ── 메인 ─────────────────────────────────────────────────
def main():
    base_dir = get_script_dir()
    excel_path = os.path.join(base_dir, EXCEL_FILE)
    output_dir = os.path.join(base_dir, OUTPUT_DIR)

    print("=" * 55)
    print("  기업건강의학센터 업무실적 — 엑셀 변환기 (통합본)")
    print("=" * 55)

    os.makedirs(output_dir, exist_ok=True)
    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── [기존 기능] 업무 실적.xlsx ───────────────────────
    knx_data, req_data = [], []
    if os.path.exists(excel_path):
        print(f"\n[기존] 엑셀 파일 읽는 중... {EXCEL_FILE}")
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
            sheet_names = wb.sheetnames
            print(f"       시트 목록: {sheet_names}")

            if len(sheet_names) >= 1:
                knx_data = parse_knx(wb[sheet_names[0]])
                print(f"       → KNX {len(knx_data)}건 파싱 완료")
            if len(sheet_names) >= 2:
                req_data = parse_req(wb[sheet_names[1]])
                print(f"       → 자료요청서 {len(req_data)}건 파싱 완료")
            wb.close()
        except Exception as e:
            print(f"       [오류] {EXCEL_FILE} 처리 실패: {e}")
    else:
        print(f"\n[기존] '{EXCEL_FILE}' 파일이 없어 건너뜁니다.")

    write_js(os.path.join(output_dir, 'knx.js'), 'KNX_DATA', knx_data, generated_at)
    write_js(os.path.join(output_dir, 'req.js'), 'REQ_DATA', req_data, generated_at)
    print(f"       ✓ data/knx.js  ({len(knx_data)}건)")
    print(f"       ✓ data/req.js  ({len(req_data)}건)")

    # ── [신규 기능] 거래처별 사후관리 집계_*.xlsx ────────
    print(f"\n[신규] '{HGC_PATTERN}' 패턴 파일 탐색 중...")
    hgc_data = process_hgc(base_dir)
    write_js(os.path.join(output_dir, 'hgc.js'), 'HGC_DATA', hgc_data, generated_at)
    print(f"       ✓ data/hgc.js  ({len(hgc_data)}개 기간 그룹)")

    print()
    print("=" * 55)
    print("  변환 완료!")
    print("=" * 55)
    print(f"  KNX 발송      : {len(knx_data):>5}건")
    print(f"  자료요청서    : {len(req_data):>5}건")
    print(f"  사후관리 그룹 : {len(hgc_data):>5}개")
    print()
    print("  다음 단계:")
    print("  1. data/ 폴더를 NAS의 index.html과 같은 위치에 복사")
    print("  2. 브라우저에서 index.html 새로고침")
    print("=" * 55)
    input("\n엔터를 누르면 종료합니다...")


if __name__ == '__main__':
    try:
        main()
    except Exception:
        print("\n[예상치 못한 오류]")
        traceback.print_exc()
        input("\n엔터를 누르면 종료합니다...")