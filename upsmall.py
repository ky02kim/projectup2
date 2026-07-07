import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Border, Side, Alignment)

class HGCReader:

    def __init__(self, filename):

        self.filename = filename

        self.data = []

    def load(self):

        path = Path(self.filename)

        if not path.exists():
            raise FileNotFoundError(self.filename)

        text = path.read_text(
            encoding="utf-8"
        ).strip()

        # js 변수 형태 지원
        if text.startswith("const"):
            idx = text.find("[")
            text = text[idx:]

        self.data = json.loads(text)

        return self.data

    @property
    def period(self):

        if len(self.data) == 0:
            return ""

        s = datetime.strptime(
            self.data[0]["periodStart"],
            "%Y-%m-%d"
        )

        e = datetime.strptime(
            self.data[0]["periodEnd"],
            "%Y-%m-%d"
        )

        return f"{s.month}/{s.day}~{e.month}/{e.day}"

    @property
    def file_date(self):

        if len(self.data) == 0:
            return ""

        return self.data[0]["fileDate"]

    def rows(self):

        for row in self.data:

            yield {

                "status":
                    row.get("status", ""),

                "jongYe":
                    row.get("jongYe", ""),

                "workplace":
                    row.get("workplace", ""),

                "report":
                    row.get("report", False),

                "items":
                    row.get(
                        "items",
                        [False, False, False, False]
                    ),

                "receive":
                    row.get("receive", 0)

            }

class ReportExcel:

    def __init__(self, period):

        self.period = period

        self.wb = Workbook()

        self.ws = self.wb.active

        self.ws.title = "발송현황"

        self.row = 2

        self._make_style()

        self._make_header()


    # --------------------------------------------------
    # 스타일
    # --------------------------------------------------

    def _make_style(self):

        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor="000000"
        )

        self.header_font = Font(
            name="맑은 고딕",
            bold=True,
            color="FFFFFF",
            size=11
        )

        self.body_font = Font(
            name="맑은 고딕",
            size=10
        )

        thin = Side(style="thin")

        self.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        self.center = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


    # --------------------------------------------------
    # 헤더
    # --------------------------------------------------

    def _make_header(self):

        headers = [
            "발송항목",
            "대상기간",
            "대상 사업장",
            "진행상태",
            "Item1",
            "Item2",
            "Item3",
            "Item4"
        ]

        for col, title in enumerate(headers, 1):

            cell = self.ws.cell(
                row=1,
                column=col
            )

            cell.value = title
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center
ITEM_LABELS = [
    "Item1",
    "Item2",
    "Item3",
    "Item4",
]


class Count:

    def __init__(self):

        self.target = {
            "종건": 0,
            "예건": 0
        }

        self.complete = {
            "종건": 0,
            "예건": 0
        }

    def add(self, status, jongye):

        if jongye not in ("종건", "예건"):
            return

        if status == "target":
            self.target[jongye] += 1

        elif status == "complete":
            self.complete[jongye] += 1

    def text(self, jongye):

        t = self.target[jongye]
        c = self.complete[jongye]

        if c == 0:
            return str(t)

        return f"{c}/{t}"

    @property
    def completed(self):

        return (
            self.complete["종건"] > 0
            or
            self.complete["예건"] > 0
        )


class ReportCounter:

    def __init__(self):

        self.after = Count()

        self.brain = Count()

        self.items = [
            Count(),
            Count(),
            Count(),
            Count()
        ]

    def add(self, row):

        status = row["status"]
        jongye = row["jongYe"]

        report = row["report"]

        items = row["items"]

        # --------------------
        # 사후
        # --------------------

        if report:
            self.after.add(
                status,
                jongye
            )

        # --------------------
        # 뇌심직무
        # --------------------

        if len(items) >= 2:

            if items[0] or items[1]:

                self.brain.add(
                    status,
                    jongye
                )

        # --------------------
        # Item1~4
        # --------------------

        for i in range(4):

            if i >= len(items):
                continue

            if items[i]:

                self.items[i].add(
                    status,
                    jongye
                )

    def load(self, rows):

        for row in rows:
            self.add(row)

        return self

    def rows(self):

        yield {
            "name": "사후",
            "count": self.after
        }

        yield {
            "name": "뇌심직무",
            "count": self.brain
        }

    def item_rows(self):

        for label, counter in zip(
            ITEM_LABELS,
            self.items
        ):

            yield {
                "name": label,
                "count": counter
            }

class ReportExcel:

    def __init__(self, period):

        self.period = period

        self.wb = Workbook()

        self.ws = self.wb.active

        self.ws.title = "발송현황"

        self.row = 2

        self._make_style()

        self._make_header()


    # --------------------------------------------------
    # 스타일
    # --------------------------------------------------

    def _make_style(self):

        self.header_fill = PatternFill(
            fill_type="solid",
            fgColor="000000"
        )

        self.header_font = Font(
            name="맑은 고딕",
            bold=True,
            color="FFFFFF",
            size=11
        )

        self.body_font = Font(
            name="맑은 고딕",
            size=10
        )

        thin = Side(style="thin")

        self.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        self.center = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


    # --------------------------------------------------
    # 헤더
    # --------------------------------------------------

    def _make_header(self):

        headers = [
            "발송항목",
            "대상기간",
            "대상 사업장",
            "진행상태",
            "Item1",
            "Item2",
            "Item3",
            "Item4"
        ]

        for col, title in enumerate(headers, 1):

            cell = self.ws.cell(
                row=1,
                column=col
            )

            cell.value = title
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center
# --------------------------------------------------
    # count 표시
    # --------------------------------------------------

    def count_text(self, counter, jongye):

        target = counter.target[jongye]
        complete = counter.complete[jongye]

        if complete == 0:
            return str(target)

        return f"{complete}/{target}"


    # --------------------------------------------------
    # 완료 여부
    # --------------------------------------------------

    def status_text(self, counter):

        if counter.completed:
            return "발송 완료"

        return "발송 예정"


    # --------------------------------------------------
    # 셀 스타일
    # --------------------------------------------------

    def style_cell(self, cell):

        cell.font = self.body_font
        cell.border = self.border
        cell.alignment = self.center


    # --------------------------------------------------
    # 열 너비
    # --------------------------------------------------

    def set_width(self):

        width = {
            "A":15,
            "B":18,
            "C":20,
            "D":14,
            "E":12,
            "F":12,
            "G":12,
            "H":12
        }

        for col, w in width.items():

            self.ws.column_dimensions[col].width = w


    # --------------------------------------------------
    # 행 높이
    # --------------------------------------------------

    def set_height(self):

        for r in range(1, self.ws.max_row + 1):

            self.ws.row_dimensions[r].height = 22
    
if __name__ == "__main__":
    reader = HGCReader("hgc.js")
    reader.load()

    counter = ReportCounter()
    counter.load(reader.rows())
    excel=ReportExcel(reader.period)
    excel.write(counter)
    excel.save()

    print("발송현황.xlsx 생성 완료")
