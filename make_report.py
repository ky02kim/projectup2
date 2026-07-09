"""
hgc.js 읽기 -> 사후/뇌심직무/Item1~4 집계 -> 발송현황.xlsx 생성

사용법:
    python3 make_report.py data/hgc.js [출력경로.xlsx]
    (출력경로 생략 시 ./발송현황.xlsx 로 저장)
"""

import sys
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 발송항목 정의: (표시 이름, 레코드 필터 함수)
CATEGORIES = [
    ("사후", lambda r: True),
    ("뇌심 직무", lambda r: r["items"][0] or r["items"][1]),
    ("Item1", lambda r: r["items"][0]),
    ("Item2", lambda r: r["items"][1]),
    ("Item3", lambda r: r["items"][2]),
    ("Item4", lambda r: r["items"][3]),
]


def load_records_from_js(path):
    """'const 변수 = [ {...} ];' 형태 js 파일 또는 순수 JSON 배열 파일을 읽어 레코드 리스트 반환"""
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()

    start, end = text.find("["), text.rfind("]")
    if start == -1 or end == -1:
        raise ValueError("배열 형태([...])를 파일에서 찾지 못했습니다.")

    array_text = re.sub(r",\s*([\]}])", r"\1", text[start:end + 1])  # trailing comma 제거
    return json.loads(array_text)


def _fmt_period(records):
    if not records:
        return ""
    d1 = datetime.strptime(records[0]["periodStart"], "%Y-%m-%d")
    d2 = datetime.strptime(records[0]["periodEnd"], "%Y-%m-%d")
    return f"{d1.month}/{d1.day}~{d2.month}/{d2.day}"


def _count_by_jongye(records):
    """
    total(대상)은 status=='target' 레코드 수만 세고,
    complete(완료)는 status=='complete' 레코드 수만 센다.
    같은 사업장이 complete/target 두 줄로 각각 들어있는 구조를 전제로 함.
    """
    counts = {"종건": {"complete": 0, "total": 0}, "예건": {"complete": 0, "total": 0}}
    for r in records:
        jy = r["jongYe"]
        if r["status"] == "target":
            counts[jy]["total"] += 1
        elif r["status"] == "complete":
            counts[jy]["complete"] += 1
    return counts


def aggregate(records):
    """CATEGORIES 별로 종건/예건 완료·대상 건수를 집계해서 표 행 리스트로 반환"""
    rows = []
    for label, cond in CATEGORIES:
        filtered = [r for r in records if cond(r)]
        counts = _count_by_jongye(filtered)
        has_complete = any(r["status"] == "complete" for r in filtered)

        lines = []
        for jy in ("종건", "예건"):
            c = counts[jy]
            if c["total"] == 0:
                continue
            lines.append(f"{jy} {c['complete']}/{c['total']}")

        status = "-" if not filtered else ("발송완료" if has_complete else "발송예정")

        rows.append({
            "발송항목": label,
            "대상 기간": _fmt_period(filtered) or _fmt_period(records),
            "대상 사업장": lines if lines else ["-"],
            "진행 상태": status,
        })
    return rows


def group_by_period(records):
    """(periodStart, periodEnd) 기준으로 레코드를 묶어서 { (start, end): [records] } 반환"""
    groups = {}
    for r in records:
        key = (r["periodStart"], r["periodEnd"])
        groups.setdefault(key, []).append(r)
    return groups


def _sheet_name(period_start, period_end):
    d1 = datetime.strptime(period_start, "%Y-%m-%d")
    d2 = datetime.strptime(period_end, "%Y-%m-%d")
    return f"{d1.month}.{d1.day}-{d2.month}.{d2.day}"


def _write_sheet(ws, rows):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["발송항목", "대상 기간", "대상 사업장", "진행 상태"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border

    r = 2
    for row in rows:
        entries = row["대상 사업장"]
        span = len(entries)
        start_row, end_row = r, r + span - 1

        for col, val in ((1, row["발송항목"]), (2, row["대상 기간"]), (4, row["진행 상태"])):
            ws.cell(row=start_row, column=col, value=val)
            if span > 1:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

        for i, entry in enumerate(entries):
            c = ws.cell(row=start_row + i, column=3, value=entry)
            c.alignment = center
            c.border = border

        for rr in range(start_row, end_row + 1):
            for col in (1, 2, 4):
                cell = ws.cell(row=rr, column=col)
                cell.alignment = center
                cell.border = border

        r = end_row + 1

    for i, w in enumerate([14, 16, 18, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_excel(records, path="발송현황.xlsx"):
    """기간별로 시트를 나눠서 저장. 가장 최신 기간이 Sheet1(맨 앞)이 되도록 정렬"""
    groups = group_by_period(records)
    # periodStart 기준 내림차순(최신이 먼저) 정렬
    sorted_keys = sorted(groups.keys(), key=lambda k: k[0], reverse=True)

    wb = Workbook()
    default_ws = wb.active

    for i, key in enumerate(sorted_keys):
        period_start, period_end = key
        rows = aggregate(groups[key])
        sheet_name = _sheet_name(period_start, period_end)

        ws = default_ws if i == 0 else wb.create_sheet()
        ws.title = sheet_name
        _write_sheet(ws, rows)

    wb.save(path)
    return path


if __name__ == "__main__":
    import os
    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.withdraw()

    try:
        # exe(또는 .py) 파일이 있는 폴더를 기준으로 경로를 잡음 (더블클릭 시 작업 폴더가 달라져도 안전)
        if getattr(sys, "frozen", False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))

        js_path = os.path.join(base_dir, "data", "hgc.js")

        records = load_records_from_js(js_path)

        # 파일명은 가장 최신(=periodStart가 가장 늦은) 기간 기준으로 생성
        latest_start, latest_end = max(
            {(r["periodStart"], r["periodEnd"]) for r in records}, key=lambda k: k[0]
        )
        d1 = datetime.strptime(latest_start, "%Y-%m-%d")
        d2 = datetime.strptime(latest_end, "%Y-%m-%d")
        out_path = os.path.join(base_dir, f"집계 {d1.month}.{d1.day}-{d2.month}.{d2.day}.xlsx")

        saved = write_excel(records, out_path)
        messagebox.showinfo("완료", f"엑셀 파일이 생성되었습니다:\n{saved}")
    except Exception as e:
        messagebox.showerror("오류 발생", str(e))
